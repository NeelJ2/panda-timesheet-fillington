import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from google.oauth2 import service_account
from googleapiclient.discovery import build

# Function to extract hours and position from event title
def extract_hours_and_position(title, name):
    if name in title:
        start = title.find(f"{name} (") + len(name) + 2
        end = title.find(")", start)
        info = title[start:end]
        role_code, hours = info.split(' ')
        hours = float(hours)
        role_map = {
            'M': 'Summer Manager',
            'S': 'Summer Teacher'
        }
        position = role_map.get(role_code, None)
        return hours, position
    return None, None

# Path to the credentials.json file
SERVICE_ACCOUNT_FILE = 'credentials.json'
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']

# Authenticate and construct the service
credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)

service = build('calendar', 'v3', credentials=credentials)

# Calculate the start and end of the current month dynamically
now = datetime.datetime.utcnow()
start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
first_day_next_month = (start_of_month + datetime.timedelta(days=32)).replace(day=1)
end_of_month = first_day_next_month - datetime.timedelta(seconds=1)

# Format dates for Google Calendar API
start_of_month_str = start_of_month.isoformat() + 'Z'
end_of_month_str = end_of_month.isoformat() + 'Z'

# Fetch events from the calendar for the current month
calendar_id = 'neel.joshi301@gmail.com'
events_result = service.events().list(calendarId=calendar_id, timeMin=start_of_month_str,
                                      timeMax=end_of_month_str, singleEvents=True,
                                      orderBy='startTime').execute()
events = events_result.get('items', [])

# Extract names from event titles
valid_names = set()
for event in events:
    title = event.get('summary', '')
    names_in_title = title.split('&')
    for name in names_in_title:
        name = name.strip().split('(')[0].strip()
        valid_names.add(name)

# Ask for the user's name and validate it
while True:
    user_name = input("Please enter your name (e.g., Neel J.): ").strip()
    if user_name in valid_names:
        break
    print("Name not found in the calendar events. Please try again.")

# Parse events into a list of dictionaries
events_data = []
for event in events:
    start = event['start'].get('dateTime', event['start'].get('date'))
    end = event['end'].get('dateTime', event['end'].get('date'))
    location = event.get('location', 'No location specified')
    title = event.get('summary', '')
    
    hours, position = extract_hours_and_position(title, user_name)
    if hours is not None:
        events_data.append({'start': start, 'end': end, 'location': location, 'hours': hours, 'position': position})

# Load the existing timesheet
file_path = 'Timesheet 2023.xlsx'
workbook = load_workbook(filename=file_path)
sheet = workbook.active

# Define the cell positions based on your timesheet structure
# Assuming the data starts from row 4
start_row = 4
date_col = 1
hours_col = 2
location_col = 3
position_col = 4

# Define the drop-down menu options
roles = [
    'Back Office', 'ISFT Assistant', 'ISFT Lead', 'PSS', 'Special Event', 
    'Summer Manager', 'Summer Teacher', 'Teacher - Assistant', 
    'Teacher - Lead', 'Teacher - Online Class'
]

# Add data validation for the roles column
dv = DataValidation(type="list", formula1='"{}"'.format(','.join(roles)), showDropDown=True)
sheet.add_data_validation(dv)

# Update the timesheet with events data
for i, event in enumerate(events_data):
    start_time = pd.to_datetime(event['start'])
    
    row = start_row + i
    sheet.cell(row=row, column=date_col, value=start_time.date())
    sheet.cell(row=row, column=hours_col, value=event['hours'])
    sheet.cell(row=row, column=location_col, value=event['location'])
    
    # Make sure the position is set correctly for the first row
    position_value = event['position'] if event['position'] else 'Unknown'
    position_cell = sheet.cell(row=row, column=position_col, value=position_value)
    dv.add(position_cell)

# Save the updated timesheet
mydate = datetime.datetime.now()
file_name = str(mydate.strftime("%B")) + " Timesheet.xlsx"
workbook.save(filename=file_name)

print(f"Timesheet saved as {file_name}")
