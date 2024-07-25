import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from google.oauth2 import service_account
from googleapiclient.discovery import build

def fetch_calendar_events(service, calendar_id, start_of_month_str, end_of_month_str):
    events_result = service.events().list(calendarId=calendar_id, timeMin=start_of_month_str,
                                          timeMax=end_of_month_str, singleEvents=True,
                                          orderBy='startTime').execute()
    events = events_result.get('items', [])
    print(f'Fetched events: {events}')  # Debug statement
    return events

def extract_hours_and_position(title, name):
    if name in title:
        start = title.find(f"{name} (") + len(name) + 2
        end = title.find(")", start)
        info = title[start:end]
        role_code, hours = info.split(' ')
        hours = float(hours)
        role_map = {
            'M': 'Summer Manager',
            'S': 'Summer Teacher'
        }
        position = role_map.get(role_code, None)
        return hours, position
    return None, None

def parse_events(events, user_name):
    events_data = []
    for event in events:
        start = event['start'].get('dateTime', event['start'].get('date'))
        end = event['end'].get('dateTime', event['end'].get('date'))
        location = event.get('location', 'No location specified')
        title = event.get('summary', '')

        hours, position = extract_hours_and_position(title, user_name)
        if hours is not None:
            events_data.append({'start': start, 'end': end, 'location': location, 'hours': hours, 'position': position})
    print(f'Parsed events data: {events_data}')  # Debug statement
    return events_data

def fill_timesheet_with_events(events_data, config, file_name):
    start_row = config['start_row']
    date_col = config['columns']['date_col']
    hours_col = config['columns']['hours_col']
    location_col = config['columns']['location_col']
    position_col = config['columns']['position_col']

    workbook = Workbook()
    sheet = workbook.active

    roles = [
        'Back Office', 'ISFT Assistant', 'ISFT Lead', 'PSS', 'Special Event', 
        'Summer Manager', 'Summer Teacher', 'Teacher - Assistant', 
        'Teacher - Lead', 'Teacher - Online Class'
    ]

    dv = DataValidation(type="list", formula1='"{}"'.format(','.join(roles)), showDropDown=True)
    sheet.add_data_validation(dv)

    for i, event in enumerate(events_data):
        start_time = pd.to_datetime(event['start'])

        row = start_row + i
        sheet.cell(row=row, column=date_col, value=start_time.date())
        sheet.cell(row=row, column=hours_col, value=event['hours'])
        sheet.cell(row=row, column=location_col, value=event['location'])
        position_value = event['position'] if event['position'] else 'Unknown'
        position_cell = sheet.cell(row=row, column=position_col, value=position_value)
        dv.add(position_cell)

    workbook.save(filename=file_name)
