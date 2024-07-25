from flask import Flask, request, jsonify, send_file, redirect, url_for, session, render_template
from flask_session import Session
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from scraper import fetch_calendar_events, parse_events  # Ensure these functions are correctly defined in scraper.py
import os
import json
import datetime
import pandas as pd

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['SESSION_TYPE'] = 'filesystem'
Session(app)

# Google OAuth 2.0 setup
CLIENT_SECRETS_FILE = "client_secret.json"
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']

@app.before_first_request
def clear_session():
    session.clear()

def get_redirect_uri():
    host = request.host
    scheme = request.scheme
    return f'{scheme}://{host}/oauth2callback'

@app.route('/')
def index():
    if 'credentials' not in session:
        return render_template('login.html')
    return redirect(url_for('fill_timesheet'))

@app.route('/authorize')
def authorize():
    session.clear()  # Clear the session to force re-login
    flow = Flow.from_client_secrets_file(
        CLIENT_SECRETS_FILE,
        scopes=SCOPES,
        redirect_uri=get_redirect_uri()
    )
    authorization_url, state = flow.authorization_url()
    session['state'] = state
    app.logger.debug(f'Authorization URL: {authorization_url}')
    app.logger.debug(f'State set in session: {state}')
    return redirect(authorization_url)

@app.route('/oauth2callback')
def oauth2callback():
    app.logger.debug(f'Received state: {request.args.get("state")}')
    app.logger.debug(f'Session state: {session.get("state")}')

    if 'state' not in session or session['state'] != request.args['state']:
        app.logger.error('State mismatch or missing state in session')
        return redirect(url_for('index'))

    try:
        flow = Flow.from_client_secrets_file(
            CLIENT_SECRETS_FILE,
            scopes=SCOPES,
            redirect_uri=get_redirect_uri()
        )
        flow.fetch_token(authorization_response=request.url)

        credentials = flow.credentials
        app.logger.debug(f'Credentials token: {credentials.token}')
        session['credentials'] = credentials_to_dict(credentials)
        app.logger.debug(f'Session credentials: {session["credentials"]}')
        return redirect(url_for('fill_timesheet'))
    except Exception as e:
        app.logger.error(f'Error during OAuth callback: {e}', exc_info=True)
        return jsonify({'error': f'An error occurred during the OAuth callback: {str(e)}'}), 403

@app.route('/fill-timesheet', methods=['GET', 'POST'])
def fill_timesheet():
    if 'credentials' not in session:
        app.logger.error('No credentials in session, redirecting to authorize')
        return redirect(url_for('authorize'))

    credentials = Credentials(**session['credentials'])
    service = build('calendar', 'v3', credentials=credentials)

    if request.method == 'POST':
        user_name = request.form['user_name']

        # Load configuration
        with open('config.json') as config_file:
            config = json.load(config_file)

        # Calculate the start and end of the current month dynamically
        now = datetime.datetime.utcnow()
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        first_day_next_month = (start_of_month + datetime.timedelta(days=32)).replace(day=1)
        end_of_month = first_day_next_month - datetime.timedelta(seconds=1)

        # Format dates for Google Calendar API
        start_of_month_str = start_of_month.isoformat() + 'Z'
        end_of_month_str = end_of_month.isoformat() + 'Z'

        try:
            # Fetch events from the calendar for the current month
            events = fetch_calendar_events(service, 'primary', start_of_month_str, end_of_month_str)
            app.logger.debug(f'Fetched events: {events}')

            # Parse events for the given user
            events_data = parse_events(events, user_name)
            app.logger.debug(f'Parsed events: {events_data}')

            # Define the filename for the timesheet
            mydate = datetime.datetime.now()
            file_name = str(mydate.strftime("%B")) + " Timesheet.xlsx"

            # Fill the timesheet with events data
            fill_timesheet_with_events(events_data, config, file_name)
            app.logger.debug(f'Timesheet generated: {file_name}')

            return send_file(file_name, as_attachment=True)

        except HttpError as error:
            error_message = json.loads(error.content)['error']['message']
            app.logger.error(f'Google API error: {error_message}')
            return jsonify({'error': f'An error occurred: {error_message}'}), 403

    return render_template('fill_timesheet.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

def credentials_to_dict(credentials):
    return {
        'token': credentials.token,
        'refresh_token': credentials.refresh_token,
        'token_uri': credentials.token_uri,
        'client_id': credentials.client_id,
        'client_secret': credentials.client_secret,
        'scopes': credentials.scopes
    }

def fill_timesheet_with_events(events_data, config, file_name):
    # Load the existing timesheet
    file_path = 'Timesheet 2023.xlsx'
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Define the cell positions based on your timesheet structure
    # Assuming the data starts from row 4
    start_row = 4
    date_col = 1
    hours_col = 2
    location_col = 3
    position_col = 4

    # Define the drop-down menu options
    roles = [
        'Back Office', 'ISFT Assistant', 'ISFT Lead', 'PSS', 'Special Event', 
        'Summer Manager', 'Summer Teacher', 'Teacher - Assistant', 
        'Teacher - Lead', 'Teacher - Online Class'
    ]

    # Add data validation for the roles column
    dv = DataValidation(type="list", formula1='"{}"'.format(','.join(roles)), showDropDown=True)
    sheet.add_data_validation(dv)

    # Update the timesheet with events data
    for i, event in enumerate(events_data):
        start_time = pd.to_datetime(event['start'])
        
        row = start_row + i
        sheet.cell(row=row, column=date_col, value=start_time.date())
        sheet.cell(row=row, column=hours_col, value=event['hours'])
        sheet.cell(row=row, column=location_col, value=event['location'])
        
        # Make sure the position is set correctly for the first row
        position_value = event['position'] if event['position'] else 'Unknown'
        position_cell = sheet.cell(row=row, column=position_col, value=position_value)
        dv.add(position_cell)

    # Save the updated timesheet
    workbook.save(filename=file_name)
    app.logger.debug(f"Timesheet saved as {file_name}")

if __name__ == '__main__':
    app.run(ssl_context=('cert.pem', 'key.pem'), debug=True)
